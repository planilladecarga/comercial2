#!/usr/bin/env python3
"""
=============================================================================
FRIMARAL BI — Libro Maestro de Business Intelligence
=============================================================================
Genera el libro Excel profesional con todas las hojas del proyecto BI.

Uso:
    python build_libro_maestro.py [xlsb_entrada.xlsx]

Si se proporciona un archivo XLSB/XLSX de origen, lo importa automáticamente.
Si no, genera la estructura vacía lista para recibir datos.

Autor  : FRIMARAL BI Team
Versión: 1.0.0
=============================================================================
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, Protection
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.formatting.rule import DataBarRule
from datetime import datetime, date
import os
import sys

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURACIÓN GLOBAL DE ESTILO
# ─────────────────────────────────────────────────────────────────────────────

# Paleta de colores profesional
COLOR_NAV      = "1F3864"   # Azul oscuro (navegación / headers principales)
COLOR_HEADER   = "2E75B6"   # Azul medio (encabezados de columna)
COLOR_SUBHDR   = "D6E4F0"   # Azul muy claro (sub-headers)
COLOR_ALT_ROW  = "EBF3FB"   # Azul Alternado muy suave
COLOR_ACCENT   = "E67E22"   # Naranja FRIMARAL (acentos)
COLOR_WHITE    = "FFFFFF"
COLOR_LIGHT_GRAY = "F2F2F2"
COLOR_TEXT_DARK = "1C1C1C"

# Bordes
thin  = Side(style="thin",   color="B8CCE4")
thick = Side(style="medium", color="2E75B6")

BORDER_THIN  = Border(left=thin,  right=thin,  top=thin,  bottom=thin)
BORDER_THICK = Border(left=thick, right=thick, top=thick, bottom=thick)
BORDER_NONE  = Border()

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS DE ESTILO
# ─────────────────────────────────────────────────────────────────────────────

def _make_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _make_font(bold: bool = False, size: int = 10,
               color: str = COLOR_TEXT_DARK, italic: bool = False) -> Font:
    return Font(bold=bold, size=size, color=color, italic=italic,
                name="Calibri")


def style_cell(cell, bold=False, size=10, color=COLOR_TEXT_DARK,
               bg=None, align_h="left", align_v="center",
               wrap=False, italic=False, border=None):
    cell.font      = _make_font(bold=bold, size=size, color=color, italic=italic)
    cell.alignment = Alignment(horizontal=align_h, vertical=align_v,
                               wrap_text=wrap)
    if bg:
        cell.fill = _make_fill(bg)
    cell.border = border or BORDER_THIN


def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width


def freeze_top_row(ws, row=2):
    """Congela la fila de encabezados."""
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def add_version_footer(ws, version="1.0.0"):
    """Agrega pie de versión profesional."""
    last_row = ws.max_row + 2
    ws.merge_cells(f"A{last_row}:Z{last_row}")
    cell = ws[f"A{last_row}"]
    cell.value = f"FRIMARAL BI  |  Libro Maestro v{version}  |  "
    cell.value += f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    cell.font      = _make_font(size=8, color="808080", italic=True)
    cell.alignment = Alignment(horizontal="right")


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS DE TABLA ESTRUCTURADA
# ─────────────────────────────────────────────────────────────────────────────

def create_table(ws, ref: str, name: str,
                 style: str = "TableStyleMedium2") -> Table:
    """Crea una tabla estructurada con nombre."""
    tbl = Table(displayName=name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name=style,
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True,   showColumnStripes=False)
    ws.add_table(tbl)
    return tbl


def add_header_row(ws, row: int, headers: list,
                   bg=COLOR_HEADER, fg=COLOR_WHITE,
                   bold=True, size=10):
    """Escribe una fila de encabezados con estilo."""
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        style_cell(cell, bold=bold, size=size, color=fg, bg=bg,
                   align_h="center", align_v="center",
                   border=BORDER_THIN)


# ─────────────────────────────────────────────────────────────────────────────
# HOJAS DEL LIBRO
# ─────────────────────────────────────────────────────────────────────────────

def build_00_inicio(wb: Workbook):
    """
    HOJA 00: Inicio — Tablero de navegación central.
    """
    ws = wb.active
    ws.title = "00_Inicio"

    # Ancho de columnas
    for col, w in enumerate("AABBBCCCCCCCCCCCCCCCCCCCCCCCC", 1):
        ws.column_dimensions[get_column_letter(col)].width = w.count("C") + 4

    # Banner superior
    ws.merge_cells("A1:Z1")
    ws.row_dimensions[1].height = 48
    cell = ws["A1"]
    cell.value = "📊  FRIMARAL BI  —  Libro Maestro de Business Intelligence"
    cell.font      = Font(bold=True, size=22, color=COLOR_WHITE, name="Calibri")
    cell.fill      = _make_fill(COLOR_NAV)
    cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:Z2")
    ws["A2"].value = "Validación oficial para toda aplicación futura | MGAP Uruguay"
    ws["A2"].font      = Font(size=10, color="C0C0C0", italic=True, name="Calibri")
    ws["A2"].fill      = _make_fill(COLOR_NAV)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[2].height = 18

    # Línea decorativa
    ws.merge_cells("A3:Z3")
    ws["A3"].fill = _make_fill(COLOR_ACCENT)
    ws.row_dimensions[3].height = 4

    # Índice general
    row = 5
    ws.cell(row=row, column=1, value="ÍNDICE DE HOJAS").font = Font(
        bold=True, size=13, color=COLOR_NAV, name="Calibri")
    ws.row_dimensions[row].height = 22
    row += 1

    sections = [
        ("CONFIGURACIÓN", [
            ("01_Datos_MGAP",       "01", "Datos crudos del MGAP — importación原始"),
            ("02_Datos_Normalizados","02", "Datos limpiados y estandarizados"),
            ("03_Diccionario_MGAP",  "03", "Diccionario de campos y definiciones"),
            ("04_Reglas_Negocio",    "04", "Validaciones y reglas de negocio"),
        ]),
        ("DIMENSIONES", [
            ("05_Empresas",         "05", "Catálogo de empresas"),
            ("06_Productores",      "06", "Catálogo de productores"),
            ("07_Depositos",        "07", "Catálogo de depósitos"),
            ("08_Certificadores",   "08", "Catálogo de certificadores"),
            ("09_Mercados",         "09", "Catálogo de mercados"),
            ("10_Productos",        "10", "Catálogo de productos"),
            ("11_Cortes",          "11", "Catálogo de tipos de corte"),
            ("12_Calendario",      "12", "Calendario perpetuo 2025-2030"),
        ]),
        ("ANÁLISIS", [
            ("13_Catalogo_Consultas","13", "Catálogo de consultas predefinidas"),
        ]),
        ("FUTURO (en construcción)", [
            ("20_Tablas_Dinamicas", "20", "Tablas dinámicas — en desarrollo"),
            ("30_Dashboard",        "30", "Dashboard — en desarrollo"),
            ("40_Empresas360",      "40", "Vista 360 de empresas — en desarrollo"),
            ("50_Competencia",      "50", "Análisis de competencia — en desarrollo"),
            ("60_Radar_Comercial",  "60", "Radar comercial — en desarrollo"),
        ]),
    ]

    def add_section(title, items):
        nonlocal row
        # Sección header
        ws.merge_cells(f"A{row}:F{row}")
        c = ws.cell(row=row, column=1, value=f"  {title}")
        c.font      = Font(bold=True, size=11, color=COLOR_WHITE, name="Calibri")
        c.fill      = _make_fill(COLOR_HEADER)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 20
        row += 1

        # Column headers
        for col, hdr in enumerate(["Hoja", "Código", "Descripción", "", "", ""], 1):
            c = ws.cell(row=row, column=col, value=hdr)
            style_cell(c, bold=True, size=9, color=COLOR_HEADER,
                       bg=COLOR_SUBHDR, align_h="center")
        ws.row_dimensions[row].height = 16
        row += 1

        for sheet_name, code, desc in items:
            # Hyperlink a la hoja
            if sheet_name in wb.sheetnames:
                cell = ws.cell(row=row, column=1, value=sheet_name)
                cell.hyperlink = f"#{sheet_name}!A1"
                cell.font = Font(bold=True, size=10, color="0563C1",
                                underline="single", name="Calibri")
            else:
                cell = ws.cell(row=row, column=1, value=f"{sheet_name} ⏳")
                cell.font = Font(bold=True, size=10, color="808080",
                                 italic=True, name="Calibri")

            ws.cell(row=row, column=2, value=code)
            ws.merge_cells(f"C{row}:F{row}")
            ws.cell(row=row, column=3, value=desc)

            for col in range(1, 7):
                c = ws.cell(row=row, column=col)
                style_cell(c, size=10, bg=COLOR_ALT_ROW if row % 2 == 0 else COLOR_WHITE)

            ws.row_dimensions[row].height = 18
            row += 1

        row += 1  # espacio entre secciones

    for section_title, items in sections:
        add_section(section_title, items)

    # Leyenda de colores
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws.cell(row=row, column=1, value="  LEYENDA DE ESTADO").font = Font(
        bold=True, size=11, color=COLOR_WHITE, name="Calibri")
    ws.cell(row=row, column=1).fill = _make_fill(COLOR_NAV)
    ws.row_dimensions[row].height = 20
    row += 1

    legend = [
        ("🔵 Hoja activa",         "Datos cargados y disponibles"),
        ("⏳ En construcción",    "Hoja estructurada, pendiente de datos"),
        ("🟡 Próximamente",        "Diseño definido, desarrollo futuro"),
    ]
    for emoji_label, desc in legend:
        ws.cell(row=row, column=1, value=emoji_label).font = Font(
            bold=True, size=10, name="Calibri")
        ws.merge_cells(f"B{row}:F{row}")
        ws.cell(row=row, column=2, value=desc).font = Font(size=10, name="Calibri")
        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = _make_fill(COLOR_LIGHT_GRAY)
        ws.row_dimensions[row].height = 16
        row += 1

    add_version_footer(ws)
    freeze_top_row(ws, row=0)


def build_01_datos_mgap(wb: Workbook, raw_data=None):
    """
    HOJA 01: Datos_MGAP — Importación de datos crudos del XLSB.
    """
    ws = wb.create_sheet("01_Datos_MGAP")

    # Anchos
    col_widths = [18, 25, 22, 22, 18, 18, 15, 20, 20, 15, 18, 18, 22, 22, 15]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Título
    ws.merge_cells("A1:O1")
    c = ws["A1"]
    c.value = "01 — Datos MGAP (Crudos)"
    c.font      = Font(bold=True, size=16, color=COLOR_WHITE, name="Calibri")
    c.fill      = _make_fill(COLOR_NAV)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:O2")
    ws["A2"].value = "Fuente: XLSB MGAP Uruguay — Movimientos desde 01/01/2025"
    ws["A2"].font      = Font(size=9, color="808080", italic=True, name="Calibri")
    ws["A2"].fill      = _make_fill(COLOR_SUBHDR)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    # Encabezados estándar MGAP (preparados para el XLSB real)
    headers = [
        "Nro_Establecimiento", "Nombre_Establecimiento", "Departamento",
        "Localidad", "Tipo_Establecimiento", "Nro_Productor", "Nombre_Productor",
        "RUC_Empresa", "Nombre_Empresa", "Tipo_Movimiento", "Fecha_Movimiento",
        "Nro_Gre", "Nro_Certificado", "Producto", "Kilos_Netos"
    ]
    add_header_row(ws, 3, headers)

    if raw_data:
        for r_idx, row_data in enumerate(raw_data, start=4):
            for c_idx, val in enumerate(row_data, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)
    else:
        # Placeholder rows con ejemplo
        placeholder = [
            "00001", "Estab. La Esperanza", "Canelones", "Las Piedras",
            "Productor", "00001", "Juan Pérez", "020123450016",
            "FRIMARAL S.A.", "Faena", "01/01/2025",
            "GRE-2025-00001", "CERT-2025-00001", "Carne Bovina", "500.00"
        ]
        for r in range(4, 9):
            for c, val in enumerate(placeholder, 1):
                cell = ws.cell(row=r, column=c, value=val if r == 4 else "")
                style_cell(cell, size=10,
                           bg=COLOR_ALT_ROW if r % 2 == 0 else COLOR_WHITE)
            ws.row_dimensions[r].height = 16

    # Tabla estructurada
    last_row = ws.max_row
    create_table(ws, f"A3:O{last_row}", "tbl_DatosMGAP")

    # Nota
    note_row = last_row + 2
    ws.merge_cells(f"A{note_row}:O{note_row}")
    ws.cell(row=note_row, column=1,
            value="ℹ  Los datos de esta hoja son de solo lectura — "
                  "procedentes del XLSB oficial del MGAP. "
                  "Para actualizar, ejecutar el script de importación.")
    ws.cell(row=note_row, column=1).font = Font(
        size=9, italic=True, color="808080", name="Calibri")

    add_version_footer(ws)
    freeze_top_row(ws, row=3)


def build_02_datos_normalizados(wb: Workbook):
    """
    HOJA 02: Datos Normalizados — Datos limpiados y estructurados.
    """
    ws = wb.create_sheet("02_Datos_Normalizados")

    col_widths = [18, 25, 14, 20, 20, 20, 18, 18, 15, 15, 12, 12, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A1:M1")
    c = ws["A1"]
    c.value = "02 — Datos Normalizados"
    c.font      = Font(bold=True, size=16, color=COLOR_WHITE, name="Calibri")
    c.fill      = _make_fill(COLOR_NAV)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:M2")
    ws["A2"].value = "Datos limpiados, validados y estructurados según reglas de negocio (04_Reglas_Negocio)"
    ws["A2"].font      = Font(size=9, color="808080", italic=True, name="Calibri")
    ws["A2"].fill      = _make_fill(COLOR_SUBHDR)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    headers = [
        "ID_Movimiento", "Nro_Estab", "Nombre_Estab", "Depto", "Productor_ID",
        "Empresa_ID", "Fecha_Mov", "Año", "Mes", "Trimestre",
        "Tipo_Movimiento", "Producto_ID", "Kilos_Netos"
    ]
    add_header_row(ws, 3, headers)

    # Placeholder
    for r in range(4, 9):
        for c in range(1, 14):
            style_cell(ws.cell(row=r, column=c), size=10,
                       bg=COLOR_ALT_ROW if r % 2 == 0 else COLOR_WHITE)
        ws.row_dimensions[r].height = 16

    last_row = ws.max_row
    create_table(ws, f"A3:M{last_row}", "tbl_DatosNormalizados")
    add_version_footer(ws)
    freeze_top_row(ws, row=3)


def build_03_diccionario_mgap(wb: Workbook):
    """
    HOJA 03: Diccionario MGAP — Definición de cada campo.
    """
    ws = wb.create_sheet("03_Diccionario_MGAP")

    for i, w in enumerate([28, 20, 14, 14, 40, 30], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "03 — Diccionario de Datos MGAP"
    c.font      = Font(bold=True, size=16, color=COLOR_WHITE, name="Calibri")
    c.fill      = _make_fill(COLOR_NAV)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:F2")
    ws["A2"].value = "Glosario oficial de campos — Fuente: MGAP Uruguay"
    ws["A2"].font      = Font(size=9, color="808080", italic=True, name="Calibri")
    ws["A2"].fill      = _make_fill(COLOR_SUBHDR)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    headers = ["Campo", "Tipo_Dato", "Longitud", "Obligatorio", "Descripción", "Observaciones"]
    add_header_row(ws, 3, headers)

    fields = [
        ("Nro_Establecimiento",  "Texto",    "5",   "Sí",
         "Número único de identificación del establecimiento","Formato: 00000"),
        ("Nombre_Establecimiento","Texto",   "100",  "Sí",
         "Denominación oficial del establecimiento",""),
        ("Departamento",         "Texto",    "30",  "Sí",
         "Departamento geográfico de Uruguay",""),
        ("Localidad",            "Texto",    "50",  "No",
         "Localidad específica dentro del departamento",""),
        ("Tipo_Establecimiento",  "Texto",    "20",  "Sí",
         "Clasificación: Productor / Certificador / Ambos",""),
        ("Nro_Productor",        "Texto",    "10",  "Sí",
         "Número de registro del productor",""),
        ("Nombre_Productor",      "Texto",    "100", "Sí",
         "Nombre completo del productor","Sin abreviaturas"),
        ("RUC_Empresa",          "Texto",    "14",  "Sí",
         "RUC de la empresa (12 o 14 dígitos)",""),
        ("Nombre_Empresa",        "Texto",   "100", "Sí",
         "Razón social de la empresa",""),
        ("Tipo_Movimiento",      "Texto",    "20",  "Sí",
         "Clasificación: Faena / Certificación / Despacho / others",""),
        ("Fecha_Movimiento",     "Fecha",    "10",  "Sí",
         "Fecha del movimiento (dd/mm/yyyy)","Desde 01/01/2025"),
        ("Nro_Gre",              "Texto",    "20",  "No",
         "Número de Guía de Remisión Electrónica","NULL si no aplica"),
        ("Nro_Certificado",      "Texto",    "20",  "No",
         "Número de certificado sanitario","NULL si no aplica"),
        ("Producto",             "Texto",    "50",  "Sí",
         "Descripción del producto","Carne Bovina, Ovina, others"),
        ("Kilos_Netos",          "Decimal",  "12,2","Sí",
         "Peso neto en kilogramos","Sin decimales多余的"),
    ]

    for r_idx, row_data in enumerate(fields, start=4):
        for c_idx, val in enumerate(row_data, start=1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            style_cell(c, size=10,
                       bg=COLOR_ALT_ROW if r_idx % 2 == 0 else COLOR_WHITE,
                       wrap=(c_idx in [5, 6]))
        ws.row_dimensions[r_idx].height = 18

    last_row = ws.max_row
    create_table(ws, f"A3:F{last_row}", "tbl_DiccionarioMGAP")
    add_version_footer(ws)
    freeze_top_row(ws, row=3)


def build_04_reglas_negocio(wb: Workbook):
    """
    HOJA 04: Reglas de Negocio — Validaciones y constraints.
    """
    ws = wb.create_sheet("04_Reglas_Negocio")

    for i, w in enumerate([8, 22, 12, 18, 40], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "04 — Reglas de Negocio"
    c.font      = Font(bold=True, size=16, color=COLOR_WHITE, name="Calibri")
    c.fill      = _make_fill(COLOR_NAV)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:E2")
    ws["A2"].value = "Validaciones que deben cumplir los datos en 02_Datos_Normalizados"
    ws["A2"].font      = Font(size=9, color="808080", italic=True, name="Calibri")
    ws["A2"].fill      = _make_fill(COLOR_SUBHDR)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    headers = ["ID", "Regla", "Tipo", "Tabla_Afectada", "Descripción"]
    add_header_row(ws, 3, headers)

    rules = [
        ("RN001", "FECHA_NO_NULA",       "Validación",   "02_Datos_Normalizados",
         "Fecha_Movimiento no puede ser NULL ni anterior a 01/01/2025"),
        ("RN002", "KILOS_POSITIVOS",     "Validación",   "02_Datos_Normalizados",
         "Kilos_Netos debe ser > 0"),
        ("RN003", "RUC_FORMATO",         "Validación",   "01_Datos_MGAP",
         "RUC debe tener 12 o 14 dígitos numéricos"),
        ("RN004", "ESTAB_EXISTE",         "Referencial",  "02_Datos_Normalizados",
         "Nro_Estab debe existir en tabla 05_Empresas"),
        ("RN005", "PRODUCTOR_EXISTE",     "Referencial",  "02_Datos_Normalizados",
         "Productor_ID debe existir en tabla 06_Productores"),
        ("RN006", "EMPRESA_EXISTE",       "Referencial",  "02_Datos_Normalizados",
         "Empresa_ID debe existir en tabla 05_Empresas"),
        ("RN007", "PRODUCTO_EXISTE",      "Referencial",  "02_Datos_Normalizados",
         "Producto_ID debe existir en tabla 10_Productos"),
        ("RN008", "DEPTO_VALIDO",         "Dominio",      "02_Datos_Normalizados",
         "Departamento debe ser uno de los 19 deptos de Uruguay"),
        ("RN009", "MES_CALCULADO",        "Automático",   "02_Datos_Normalizados",
         "Mes se calcula automáticamente a partir de Fecha_Mov"),
        ("RN010", "TRIMESTRE_CALCULADO",  "Automático",   "02_Datos_Normalizados",
         "Trimestre se calcula automáticamente a partir de Fecha_Mov"),
        ("RN011", "DUPLICADO_GRE",        "Validación",   "01_Datos_MGAP",
         "No puede haber dos registros con el mismo Nro_Gre"),
        ("RN012", "CORTE_VALIDO",         "Dominio",      "02_Datos_Normalizados",
         "Tipo de corte debe existir en catálogo 11_Cortes"),
    ]

    for r_idx, row_data in enumerate(rules, start=4):
        for c_idx, val in enumerate(row_data, start=1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            bg = COLOR_ALT_ROW if r_idx % 2 == 0 else COLOR_WHITE
            wrap = c_idx == 5
            style_cell(c, size=10, bg=bg, wrap=wrap)
        ws.row_dimensions[r_idx].height = 18

    last_row = ws.max_row
    create_table(ws, f"A3:E{last_row}", "tbl_ReglasNegocio")
    add_version_footer(ws)
    freeze_top_row(ws, row=3)


def build_dimension_sheet(wb: Workbook, sheet_name: str, code: str,
                            headers: list, sample_data: list,
                            id_col: int = 1):
    """
    Generador de hojas de dimensión (05-11) con estructura idéntica.
    """
    ws = wb.create_sheet(sheet_name)

    col_widths = [18] * len(headers)
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Título
    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    c = ws["A1"]
    c.value = f"{code} — {sheet_name.replace('0'+code.split('_')[0], '').replace('_',' ')}"
    c.font      = Font(bold=True, size=16, color=COLOR_WHITE, name="Calibri")
    c.fill      = _make_fill(COLOR_NAV)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    # Subtítulo
    ws.merge_cells(f"A2:{get_column_letter(len(headers))}2")
    ws["A2"].value = f"Catálogo maestro — Tabla: {sheet_name}"
    ws["A2"].font      = Font(size=9, color="808080", italic=True, name="Calibri")
    ws["A2"].fill      = _make_fill(COLOR_SUBHDR)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    add_header_row(ws, 3, headers)

    for r_idx, row_data in enumerate(sample_data, start=4):
        for c_idx, val in enumerate(row_data, start=1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            bg = COLOR_ALT_ROW if r_idx % 2 == 0 else COLOR_WHITE
            style_cell(c, size=10, bg=bg)
        ws.row_dimensions[r_idx].height = 16

    # Placeholders vacíos hasta row 14
    empty_row = 4 + len(sample_data)
    for r in range(empty_row, empty_row + 5):
        for c in range(1, len(headers) + 1):
            style_cell(ws.cell(row=r, column=c), size=10,
                       bg=COLOR_ALT_ROW if r % 2 == 0 else COLOR_WHITE)
        ws.row_dimensions[r].height = 16

    last_row = ws.max_row
    tbl_name = f"tbl_{sheet_name[3:]}"
    create_table(ws, f"A3:{get_column_letter(len(headers))}{last_row}", tbl_name)
    add_version_footer(ws)
    freeze_top_row(ws, row=3)
    return ws


def build_05_empresas(wb: Workbook):
    headers = ["ID_Empresa", "RUC", "Nombre_Empresa", "Nombre_Fantasia",
               "Tipo_Empresa", "Actividad", "Departamento", "Fecha_Inicio"]
    sample = [
        (1, "020123450016", "FRIMARAL S.A.", "FRIMARAL", "Matadero",
         "Frigorífico", "Montevideo", "01/01/2010"),
        (2, "020234560017", "CUTCSA ALIM. S.A.", "CUTCSA", "Distribuidor",
         "Distribución", "Montevideo", "15/03/2015"),
        (3, "020345670018", "RUFFONI S.A.", "Ruffoni", "Minorista",
         "Carnicería", "Canelones", "01/07/2018"),
    ]
    return build_dimension_sheet(wb, "05_Empresas", "05", headers, sample)


def build_06_productores(wb: Workbook):
    headers = ["ID_Productor", "Nro_Productor", "Nombre_Productor",
               "Documento", "Departamento", "Tipo_Productor", "Activo"]
    sample = [
        (1, "00001", "Juan Pérez", "1.234.567-8", "Canelones", "Primario", "Sí"),
        (2, "00002", "María González", "2.345.678-9", "Soriano", "Primario", "Sí"),
        (3, "00003", "Carlos Rodríguez", "3.456.789-0", "Durazno", "Primario", "Sí"),
    ]
    return build_dimension_sheet(wb, "06_Productores", "06", headers, sample)


def build_07_depositos(wb: Workbook):
    headers = ["ID_Deposito", "Nro_Deposito", "Nombre_Deposito",
               "Tipo_Deposito", "Empresa_ID", "Departamento", "Capacidad_m3", "Activo"]
    sample = [
        (1, "DEP001", "Depósito Central Montevideo", "Refrigerado", 1,
         "Montevideo", "5000", "Sí"),
        (2, "DEP002", "Cámara Fría Pecuaria", "Refrigerado", 1,
         "Montevideo", "2000", "Sí"),
        (3, "DEP003", "Depósito Las Piedras", "Seco", 2,
         "Canelones", "1500", "Sí"),
    ]
    return build_dimension_sheet(wb, "07_Depositos", "07", headers, sample)


def build_08_certificadores(wb: Workbook):
    headers = ["ID_Certificador", "Nro_Certificador", "Nombre_Certificador",
               "Tipo_Certificadora", "Acreditacion", "Vigencia_Desde", "Activo"]
    sample = [
        (1, "CERT001", "Instituto Nacional de Calidad (INAC)", "Oficial",
         "ISO 17020", "01/01/2020", "Sí"),
        (2, "CERT002", "SGS Uruguay", "Privada", "ISO 17065",
         "01/01/2021", "Sí"),
        (3, "CERT003", "Bureau Veritas Uruguay", "Privada", "ISO 17065",
         "01/06/2021", "Sí"),
    ]
    return build_dimension_sheet(wb, "08_Certificadores", "08", headers, sample)


def build_09_mercados(wb: Workbook):
    headers = ["ID_Mercado", "Codigo_Mercado", "Nombre_Mercado",
               "Tipo_Mercado", "Pais", "Region", "Activo"]
    sample = [
        (1, "UY-MVD", "Mercado Interno Montevideo", "Nacional", "Uruguay",
         "Sur", "Sí"),
        (2, "UY-LIT", "Mercado de Litoral Norte", "Nacional", "Uruguay",
         "Litoral Norte", "Sí"),
        (3, "BR-SUL", "Mercado Rio Grande do Sul", "Exportación", "Brasil",
         "Sur", "Sí"),
        (4, "AR-BUE", "Mercado de Buenos Aires", "Exportación", "Argentina",
         "Centro", "Sí"),
    ]
    return build_dimension_sheet(wb, "09_Mercados", "09", headers, sample)


def build_10_productos(wb: Workbook):
    headers = ["ID_Producto", "Codigo_SAG", "Nombre_Producto",
               "Categoria", "Unidad_Medida", "Fraccional", "Activo"]
    sample = [
        (1, "CB-001", "Carne Bovina Fresca", "Bovino", "Kg", "Sí", "Sí"),
        (2, "CO-001", "Carne Ovina Fresca",   "Ovino",  "Kg", "Sí", "Sí"),
        (3, "CB-002", "Carne Bovina Congelada","Bovino", "Kg", "Sí", "Sí"),
        (4, "PB-001", "Achuras Bovinas",      "Bovino", "Kg", "No",  "Sí"),
        (5, "CU-001", "Cueros Bovinos",       "Subprod.","Kg", "No",  "Sí"),
    ]
    return build_dimension_sheet(wb, "10_Productos", "10", headers, sample)


def build_11_cortes(wb: Workbook):
    headers = ["ID_Corte", "Codigo_Corte", "Nombre_Corte", "Producto_ID",
               "Categoria_Corte", "Grado", "Kilos_Promedio", "Activo"]
    sample = [
        (1, "COR-001", "Bife Ancho",        1, "Premium", "A", "3.5", "Sí"),
        (2, "COR-002", "Bife Angosto",       1, "Premium", "A", "2.0", "Sí"),
        (3, "COR-003", "Tira de Asado",     1, "Standard","B", "1.5", "Sí"),
        (4, "COR-004", "Vacío",              1, "Standard","B", "2.5", "Sí"),
        (5, "COR-005", "Entraña",            1, "Premium", "A", "0.8", "Sí"),
        (6, "COR-006", "Osobuco",            1, "Standard","C", "1.2", "Sí"),
    ]
    return build_dimension_sheet(wb, "11_Cortes", "11", headers, sample)


def build_12_calendario(wb: Workbook):
    """
    HOJA 12: Calendario perpetuo 2025-2030 con todas las fechas.
    """
    ws = wb.create_sheet("12_Calendario")

    for i, w in enumerate([14, 12, 12, 12, 12, 12, 12, 15, 12, 15, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A1:K1")
    c = ws["A1"]
    c.value = "12 — Calendario Perpetuo 2025–2030"
    c.font      = Font(bold=True, size=16, color=COLOR_WHITE, name="Calibri")
    c.fill      = _make_fill(COLOR_NAV)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:K2")
    ws["A2"].value = "Calendario maestro para análisis temporal — generado automáticamente"
    ws["A2"].font      = Font(size=9, color="808080", italic=True, name="Calibri")
    ws["A2"].fill      = _make_fill(COLOR_SUBHDR)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    headers = [
        "Fecha", "Año", "Mes", "Nombre_Mes", "Trimestre",
        "Semestre", "Semana_ISO", "Dia_Ano",
        "Dia_Semana", "Es_Laboral", "Festivo_UY"
    ]
    add_header_row(ws, 3, headers)

    import datetime

    def trimester(m):
        return (m - 1) // 3 + 1

    def week_of_year(d):
        return d.isocalendar()[1]

    def day_of_year(d):
        return (d - datetime.date(d.year, 1, 1)).days + 1

    days_map = {1: "Lunes", 2: "Martes", 3: "Miércoles",
                4: "Jueves",  5: "Viernes", 6: "Sábado", 7: "Domingo"}

    uy_holidays = {
        (1, 1): "Año Nuevo",
        (1, 6): "Día de los Reyes",
        (4, 17): "Desembarco de los 33 Orientales",
        (5, 1): "Día del Trabajo",
        (5, 18): "Batalla de Las Piedras",
        (6, 19): "Día de Carlos Aniani",
        (7, 18): "Jura de la Constitución",
        (8, 25): "Independencia",
        (11, 2): "Día de los Difuntos",
        (12, 25): "Navidad",
    }

    row = 4
    for year in range(2025, 2031):
        for month in range(1, 13):
            # Último día del mes
            if month == 12:
                last_day = datetime.date(year, 12, 31)
            else:
                last_day = datetime.date(year, month + 1, 1) - datetime.timedelta(days=1)

            for day in range(1, last_day.day + 1):
                d = datetime.date(year, month, day)
                d_weekday = d.weekday() + 1  # 1=Lunes, 7=Domingo
                is_laboral = d_weekday <= 5
                holiday = uy_holidays.get((month, day), "")

                values = [
                    d.strftime("%d/%m/%Y"), year, month,
                    d.strftime("%B"), trimester(month),
                    1 if month <= 6 else 2,
                    week_of_year(d), day_of_year(d),
                    days_map[d_weekday],
                    "Sí" if is_laboral else "No",
                    holiday,
                ]
                for c_idx, val in enumerate(values, 1):
                    cell = ws.cell(row=row, column=c_idx, value=val)
                    bg = COLOR_ALT_ROW if row % 2 == 0 else COLOR_WHITE
                    if not is_laboral and c_idx == 10:
                        bg = "FFF3CD"
                    if holiday and c_idx == 11:
                        bg = "FFE0B2"
                    style_cell(cell, size=9, bg=bg,
                               align_h="center" if c_idx > 1 else "left")
                ws.row_dimensions[row].height = 15
                row += 1

    last_row = row - 1
    create_table(ws, f"A3:K{last_row}", "tbl_Calendario")
    add_version_footer(ws)
    freeze_top_row(ws, row=3)


def build_13_catalogo_consultas(wb: Workbook):
    """
    HOJA 13: Catálogo de Consultas predefinidas.
    """
    ws = wb.create_sheet("13_Catalogo_Consultas")

    for i, w in enumerate([8, 30, 35, 25, 15, 40], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "13 — Catálogo de Consultas Predefinidas"
    c.font      = Font(bold=True, size=16, color=COLOR_WHITE, name="Calibri")
    c.fill      = _make_fill(COLOR_NAV)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:F2")
    ws["A2"].value = "Repositorio de consultas SQL / fórmulas de análisis — copiar y adaptar"
    ws["A2"].font      = Font(size=9, color="808080", italic=True, name="Calibri")
    ws["A2"].fill      = _make_fill(COLOR_SUBHDR)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    headers = ["ID", "Nombre_Consulta", "Descripción", "Tablas_Involucradas",
               "Complejidad", "SQL_Patron"]
    add_header_row(ws, 3, headers)

    queries = [
        ("Q001", "Movimientos por Establecimiento",
         "Lista todos los movimientos de un establecimiento en un período",
         "02_Datos_Normalizados, 05_Empresas",
         "Baja",
         "SELECT * FROM tbl_DatosNormalizados WHERE Nro_Estab = ?"),
        ("Q002", "Volumen por Producto y Mes",
         "Resumen de kilos netos por producto agrupado por mes",
         "02_Datos_Normalizados, 10_Productos",
         "Baja",
         "SELECT Mes, Producto_ID, SUM(Kilos_Netos) FROM tbl_DatosNormalizados GROUP BY Mes, Producto_ID"),
        ("Q003", "Productores Activos por Departamento",
         "Cantidad de productores únicos por departamento",
         "06_Productores",
         "Baja",
         "SELECT Departamento, COUNT(*) FROM tbl_Productores WHERE Activo = 'Sí' GROUP BY Departamento"),
        ("Q004", "Ranking de Empresas por Volumen",
         "Top 10 empresas por volumen total de movimiento",
         "02_Datos_Normalizados, 05_Empresas",
         "Media",
         "SELECT Nombre_Empresa, SUM(Kilos_Netos) as Total FROM tbl_DatosNormalizados GROUP BY Empresa_ID ORDER BY Total DESC LIMIT 10"),
        ("Q005", "Detalle de GRE por Fecha",
         "Buscar guía de remisión electrónica por rango de fechas",
         "01_Datos_MGAP",
         "Baja",
         "SELECT * FROM tbl_DatosMGAP WHERE Fecha_Movimiento BETWEEN ? AND ?"),
        ("Q006", "Análisis Trimestral por Mercado",
         "Volumen y count de movimientos por mercado y trimestre",
         "02_Datos_Normalizados, 09_Mercados",
         "Media",
         "SELECT Trimestre, Mercado_ID, SUM(Kilos_Netos) FROM tbl_DatosNormalizados GROUP BY Trimestre, Mercado_ID"),
        ("Q007", "Establecimientos Sin Movimiento",
         "Productores o depósitos sin movimiento en los últimos N días",
         "02_Datos_Normalizados, 06_Productores",
         "Alta",
         "SELECT * FROM tbl_Productores WHERE ID_Productor NOT IN (SELECT Productor_ID FROM tbl_DatosNormalizados WHERE Fecha_Mov >= DATEADD('day', -?, CURRENT_DATE))"),
    ]

    for r_idx, row_data in enumerate(queries, start=4):
        for c_idx, val in enumerate(row_data, start=1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            bg = COLOR_ALT_ROW if r_idx % 2 == 0 else COLOR_WHITE
            style_cell(c, size=10, bg=bg, wrap=c_idx in [3, 6])
        ws.row_dimensions[r_idx].height = 20

    last_row = ws.max_row
    create_table(ws, f"A3:F{last_row}", "tbl_CatalogoConsultas")
    add_version_footer(ws)
    freeze_top_row(ws, row=3)


def build_placeholder_sheet(wb: Workbook, sheet_name: str, code: str,
                             description: str):
    """
    Hoja placeholder para secciones futuras (20, 30, 40, 50, 60).
    """
    ws = wb.create_sheet(sheet_name)

    for i in range(1, 9):
        ws.column_dimensions[get_column_letter(i)].width = 18

    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = f"{code} — {sheet_name[3:].replace('_',' ')}"
    c.font      = Font(bold=True, size=16, color=COLOR_WHITE, name="Calibri")
    c.fill      = _make_fill(COLOR_NAV)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:H2")
    ws["A2"].value = description
    ws["A2"].font      = Font(size=10, italic=True, color="808080", name="Calibri")
    ws["A2"].fill      = _make_fill(COLOR_SUBHDR)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # Aviso de desarrollo futuro
    ws.merge_cells("A4:H12")
    c = ws["A4"]
    c.value = (
        "⏳  SECCIÓN EN CONSTRUCCIÓN\n\n"
        "Esta hoja será desarrollada en una fase posterior del proyecto.\n"
        "La estructura y objetivos están definidos en el plan de trabajo.\n\n"
        "Consulte el documento de especificación para más detalles."
    )
    c.font      = Font(size=13, color="808080", italic=True, name="Calibri")
    c.fill      = _make_fill(COLOR_LIGHT_GRAY)
    c.alignment = Alignment(horizontal="center", vertical="center",
                            wrap_text=True)
    ws.row_dimensions[4].height = 30
    for r in range(5, 13):
        ws.row_dimensions[r].height = 20
        for col in range(1, 9):
            ws.cell(row=r, column=col).fill = _make_fill(COLOR_LIGHT_GRAY)

    add_version_footer(ws)


# ─────────────────────────────────────────────────────────────────────────────
# BUILDER PRINCIPAL
# ─────────────────────────────────────────────────────────────────────────────

def build_libro_maestro(xlsb_path: str = None, output_path: str = None):
    """
    Construye el libro maestro completo.

    Args:
        xlsb_path   : Ruta al archivo XLSB del MGAP (opcional).
        output_path : Ruta de salida del XLSX (opcional).
                      Por defecto: ./FRIMARAL_BI_Libro_Maestro_v1_0_0.xlsx
    """
    print("🏗️  FRIMARAL BI — Generando Libro Maestro...")

    wb = Workbook()

    # Hojas principales
    print("   00_Inicio...")
    build_00_inicio(wb)

    # Datos MGAP (con importación opcional del XLSB)
    raw_data = None
    if xlsb_path and os.path.exists(xlsb_path):
        print(f"   01_Datos_MGAP (importando {xlsb_path})...")
        try:
            import pandas as pd
            if xlsb_path.endswith(".xlsb"):
                raw_data = pd.read_excel(xlsb_path, engine="pyxlsb")
            else:
                raw_data = pd.read_excel(xlsb_path)
            raw_data = [tuple(row) for row in raw_data.values.tolist()]
        except Exception as e:
            print(f"   ⚠ No se pudo importar XLSB: {e}")
            raw_data = None
    else:
        print("   01_Datos_MGAP (estructura vacía — pendiente XLSB)")
    build_01_datos_mgap(wb, raw_data)

    print("   02_Datos_Normalizados...")
    build_02_datos_normalizados(wb)

    print("   03_Diccionario_MGAP...")
    build_03_diccionario_mgap(wb)

    print("   04_Reglas_Negocio...")
    build_04_reglas_negocio(wb)

    # Dimensiones
    print("   05_Empresas...")
    build_05_empresas(wb)

    print("   06_Productores...")
    build_06_productores(wb)

    print("   07_Depositos...")
    build_07_depositos(wb)

    print("   08_Certificadores...")
    build_08_certificadores(wb)

    print("   09_Mercados...")
    build_09_mercados(wb)

    print("   10_Productos...")
    build_10_productos(wb)

    print("   11_Cortes...")
    build_11_cortes(wb)

    print("   12_Calendario...")
    build_12_calendario(wb)

    print("   13_Catalogo_Consultas...")
    build_13_catalogo_consultas(wb)

    # Placeholders para desarrollo futuro
    print("   20_Tablas_Dinamicas (placeholder)...")
    build_placeholder_sheet(wb, "20_Tablas_Dinamicas", "20",
                            "Tablas dinámicas — fase 2")

    print("   30_Dashboard (placeholder)...")
    build_placeholder_sheet(wb, "30_Dashboard", "30",
                            "Dashboard ejecutivo — fase 2")

    print("   40_Empresas360 (placeholder)...")
    build_placeholder_sheet(wb, "40_Empresas360", "40",
                            "Vista 360 de empresas — fase 3")

    print("   50_Competencia (placeholder)...")
    build_placeholder_sheet(wb, "50_Competencia", "50",
                            "Análisis de competencia — fase 3")

    print("   60_Radar_Comercial (placeholder)...")
    build_placeholder_sheet(wb, "60_Radar_Comercial", "60",
                            "Radar comercial — fase 4")

    # Guardar
    if output_path is None:
        version = "1_0_0"
        output_path = f"FRIMARAL_BI_Libro_Maestro_v{version}.xlsx"

    wb.save(output_path)
    print(f"\n✅  Libro generado: {output_path}")
    print(f"   Hojas creadas: {len(wb.sheetnames)}")
    for name in wb.sheetnames:
        print(f"     • {name}")
    return output_path


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    xlsb = sys.argv[1] if len(sys.argv) > 1 else None
    out  = sys.argv[2] if len(sys.argv) > 2 else None
    build_libro_maestro(xlsb_path=xlsb, output_path=out)
